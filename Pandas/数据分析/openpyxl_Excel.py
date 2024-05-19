
# Python数据分析

# openpyxl库常用Excel操作

# openpyxl是一个用于处理xlsx格式Excel表格文件的第三方python库，几乎支持Excel表格的所有操作
# 基本概念：
"""
Workbook：相当于一个Excel文档，每个Workbook对象都是一个独立的Excel文件
Sheet：Excel文档中的表单，每个Excel文档中至少有一个Sheet
Cell：Excel单元格，是不可分割的基本数据存储单元
"""

# 安装：pip install openpyxl

from openpyxl import load_workbook, Workbook

# 1、写入Excel

# 1）新建
# 新建一个Excel文档（初始化）
workbook = Workbook()
# 初始化/获取一个Sheet（新创建的Excel默认自带名为Sheet的表单）
# sheet = workbook.active
# 创建一个Sheet，新建的多个Sheet默认插在后面
sheet = workbook.create_sheet("Sheet1")
# 创建一个Sheet，插入到最前面
# sheet = workbook.create_sheet("Sheet1", 0)

# 2）添加数据
# append(list)：在已有的数据后面追加写入（增量写入）

# 定义表头（插入一条数据）
sheet.append(['id', 'name', 'age', 'addr'])
# 批量插入数据
data = [
    ['001', 'Tom', 18],
    ['002', 'Jerry', 17, 'US'],
    ['003', 'Alice', 20]
]
for row in data:
    sheet.append(row)

# 保存Excel-Sheet1
workbook.save(r'C:\Users\cc\Desktop\openpyxl.xlsx')

# 单元格格式
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment

# 1）字体
# 设置A1单元格字体风格为Times New Roman，大小16，粗体、斜体，蓝色
sheet['A1'].font = Font(name='Times New Roman', size=16, bold=True, italic=True, color=colors.BLUE)
# 2）对齐方式
# 设置单元格horizontal水平和vertical垂直对齐方式，其他值：left、right
sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
# 3）行高和列宽
# 设置行高
sheet.row_dimensions[1].height = 25
# 设置列宽
sheet.column_dimensions['A'].width = 15
# 4）边框
# 设置边框四个方向的线条种类
left, right, top, bottom = [Side(style='thin', color='000000')] * 4
# 将各方向线条作为参数传入Border方法
sheet['A1'].border = Border(left=left, right=right, top=top, bottom=bottom)
# 5）合并、拆分单元格
# 合并单元格
sheet.merge_cells('A1:B2')
# 拆分单元格
sheet.unmerge_cells('A1:B2')

# 保存Excel-Sheet2
workbook.save(r'C:\Users\cc\Desktop\openpyxl.xlsx')

# 2、读取Excel

# load_workbook(Excel)
# 文件必须是xlsx格式，默认为可读可写
workbook = load_workbook(r'C:\Users\cc\Desktop\openpyxl.xlsx')
# 读取指定Sheet
sheet = workbook.get_sheet_by_name('Sheet1')

# 追加一条记录
sheet.append(['004', 'Bob', 19, 'CN'])

# 保存Excel
workbook.save(r'C:\Users\cc\Desktop\openpyxl.xlsx')

# 查询

# 获取所有Sheet表名，返回List
print(workbook.sheetnames)     # ['Sheet', 'Sheet1']

# 最大行数
print(sheet.max_row)
# 最大列数
print(sheet.max_column)

# 单元格访问
print(sheet['A1'].value)
print(sheet.cell(row=4, column=2).value)
# 访问行或列
print(sheet['A'])
print(sheet['A':'C'])
print(sheet[1])
print(sheet[1:3])

# 获取所有行或列
row_list = []
for row in sheet.iter_rows():
    row_ls = []
    col_len = len(list(sheet.iter_cols()))
    for cell in row:
        col_len -= 1
        row_ls.append(cell.value)
        if col_len == 0:
            continue
    row_list.append(row_ls)

print(row_list)
'''
[['id', 'name', 'age', 'addr'], ['001', 'Tom', 18, None], ['002', 'Jerry', 17, 'US'], ['003', 'Alice', 20, None]]
'''

# 修改
# 修改指定值
sheet['C4'].value = 21
sheet.cell(row=4, column=4).value = 'CN'
# 保存Excel
workbook.save(r'C:\Users\cc\Desktop\openpyxl.xlsx')

# 删除
# 删除指定行或列
sheet.delete_rows(2)

# 删除Excel-Sheet
workbook.remove_sheet(sheet)
del workbook['Sheet']

# 删除操作后要保存
workbook.save(r'C:\Users\cc\Desktop\openpyxl.xlsx')

